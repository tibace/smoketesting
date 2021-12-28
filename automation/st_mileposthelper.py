import numpy as np
import xlsxwriter
import openpyxl
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook
import pandas as pd
from openpyxl.worksheet.table import TableStyleInfo, Table


pd.options.mode.chained_assignment = None  # default='warn'
import re
import os
import pyodbc
import sys
from openpyxl.styles import PatternFill
import shutil
import os
from django.conf import settings

# from datetime import datetime
# #--------------Getting Date and Time
# # datetime object containing current date and time
# now = datetime.now()
# # dd/mm/YY H:M:S
# dt_string = now.strftime("%d/%m/%Y_%H:%M:%S")
# print("date and time =", dt_string)
#
#
# source_dir = 'automate\\media'
# target_dir = 'automate\\media_history'
# file_names = os.listdir(source_dir)
# for file_name in file_names:
#     file = str(file_name)+''+str(dt_string)
#     shutil.move(os.path.join(source_dir, file), target_dir)



arr = os.listdir('automate\\media')
files = list(arr)
files1 = files[0]
files2 = files[1]
print(files1)
print(files2)

if files1 is not None or files1.__contains__('Design'):
    designfile=files1
    print(designfile)

if  files2 is not None or files2.__contains__('Design'):
    designfile=files2
    print(designfile)
#----------------------------------------------
if files1 is not None or files1.__contains__('Compare'):
    compfile1 = files1
    print(compfile1)

if files2 is not None or files2.__contains__('Compare'):
    compfile1=files2
    print(compfile1)


#if files1.__contains__('Design'):
file = r'C:\Users\mkanniah\automation-v1.1\automate\media\\'+designfile
file1 = r'C:\Users\mkanniah\automation-v1.1\automate\media\\'+compfile1



# converting list to string using iteration
def listToString(s):
    # initialize an empty string
    string = ""

    # traverse in the string
    for element in s:
        string += element

        # return string
    return string

def Replace_NA(df):
    total_row = len(df.index)
    total_column = len(df.columns)
    df_design_doc_new = df.iloc[:, 10:total_column + 1]

    for col in df_design_doc_new:
        i = -1
        temparray = (col[:-2])
        data = df_design_doc_new[col].values
        for row in data:
            if i < total_row - 1:
                i = i + 1
            if row == 'NA':
                df[col].values[i] = df[col].values[i].replace('NA', str(df[temparray].values[i]))

    #print(df)
    return df


def remove_dots(df):
    for col in df.columns:
        substring1 = '.'
        if col.find(substring1) != -1:
            print("Found!")
            print(col)
            temparray = (col[:-2])
            df.drop(temparray, axis=1, inplace=True)
            df.rename(
                columns={col: temparray},
                inplace=True)
            print(df)
        else:
            print("Not found!")
    return df

#--------------------Reading the Design Data ---------------------------------------------
dest_folder_design = file
df_designreport = pd.read_excel(dest_folder_design, sheet_name='Milepost Helper', na_filter=False)

print(df_designreport)

#--------------------Reading the Design Data for TrackSegment---------------------------------------------
df_designreporttrackseg = pd.read_excel(dest_folder_design, sheet_name='Track Segment', na_filter=False)
print('TrackSegment data -------------------------------------------')
#print(df_designreporttrackseg)
df_designreporttracksegment=df_designreporttrackseg[['SEGMENTID',"TRACKNAME","BEGINMILEPOST"]]
#print(df_designreporttrackseg[['SEGMENTID',"TRACKNAME","BEGINMILEPOST"]])

#----------------------below code is used find the total no of rows and columns------------
obj_workbook = openpyxl.load_workbook(file1)
print(obj_workbook.sheetnames)
obj_sheet = obj_workbook['Milepost Helper']

row_count = obj_sheet.max_row
column_count = obj_sheet.max_column
#------------------------------------------------------------------------------------------

total_column = len(df_designreport.columns)
totalchanged_column = total_column - 9
df_design_doc_new = df_designreport.iloc[:, 9:total_column + 1]
first_column_design = df_designreport.iloc[:, 0]
temarray = []


df_changed_2= pd.DataFrame()
df_removed_2= pd.DataFrame()
df_designremovechanged_2= pd.DataFrame()

for col in df_design_doc_new.columns:

    if col == 'DATASOURCEID.1' or col == 'FEATUREID.1':
        df_design_doc_new = df_designreport.iloc[:, 9:total_column + 1]
        first_column = df_design_doc_new.iloc[:, 0]
        for row in first_column:

            if str(row) == 'NA':
                Replace_NA(df_designreport)
                remove_dots(df_designreport)
                if col == 'DATASOURCEID.1':
                    df_changed = df_designreport[df_designreport["DATASOURCEID"] == 'NA']
                    df_only_changed_data = df_design_doc_new[df_design_doc_new["DATASOURCEID.1"] == 'NA']
                    df_only_changed_data = df_only_changed_data.iloc[:, 1:]

                    df_changed_2 = df_changed[
                        ['SEGMENTID',"MILEPOST", "DATASOURCEID", "FEATUREID","LATITUDE","LONGITUDE","ELEVATION"]]
                    #['SEGMENTID', "MILEPOST", "DATASOURCEID", "FEATUREID", "LATITUDE", "LONGITUDE", "ELEVATION"]]
                    df_changed_2['MILEPOST'] = df_changed['MILEPOST'].astype(str)
                    df_changed_2['DATASOURCEID'] = df_changed['DATASOURCEID'].astype(str)
                    df_changed_2['FEATUREID'] = df_changed['FEATUREID'].astype(str)
                    df_changed_2['LATITUDE'] = df_changed['LATITUDE'].astype(str)
                    df_changed_2['LONGITUDE'] = df_changed['LONGITUDE'].astype(str)
                    #df_changed_2['ELEVATION'] = df_changed['ELEVATION'].astype(str)
                    df_removed_2['SEGMENTID'] = df_removed['SEGMENTID'].astype(str)


                else:
                    df_changed = df_designreport[df_designreport["FEATUREID"] == 'NA']
                    df_only_changed_data = df_design_doc_new[df_design_doc_new["FEATUREID.1"] == 'NA']
                    df_only_changed_data_2 = df_only_changed_data.iloc[:, 1:]
                    df_changed_2 = df_changed[
                        ['SEGMENTID',"MILEPOST", "DATASOURCEID", "FEATUREID","LATITUDE","LONGITUDE"]]
                    #['SEGMENTID', "MILEPOST", "DATASOURCEID", "FEATUREID", "LATITUDE", "LONGITUDE", "ELEVATION"]]
                    df_changed_2['MILEPOST'] = df_changed['MILEPOST'].astype(str)
                    df_changed_2['DATASOURCEID'] = df_changed['DATASOURCEID'].astype(str)
                    df_changed_2['FEATUREID'] = df_changed['FEATUREID'].astype(str)
                    df_changed_2['LATITUDE'] = df_changed['LATITUDE'].astype(str)
                    df_changed_2['LONGITUDE'] = df_changed['LONGITUDE'].astype(str)
                    #df_changed_2['ELEVATION'] = df_changed['ELEVATION'].astype(str)
                    df_removed_2['SEGMENTID'] = df_removed['SEGMENTID'].astype(str)

            elif str(row) == 'REMOVE':
                # print(df_designdata)
                Replace_NA(df_designreport)
                remove_dots(df_designreport)

                if col == 'DATASOURCEID.1':
                    df_removed = df_designreport[df_designreport["DATASOURCEID"] == "REMOVE"]
                    df_removed_2 = df_removed[
                        ['SEGMENTID',"MILEPOST", "DATASOURCEID", "FEATUREID","LATITUDE","LONGITUDE"]]
                    #['SEGMENTID', "MILEPOST", "DATASOURCEID", "FEATUREID", "LATITUDE", "LONGITUDE", "ELEVATION"]]
                    df_removed_2['MILEPOST'] = df_removed['MILEPOST'].astype(str)
                    df_removed_2['DATASOURCEID'] = df_removed['DATASOURCEID'].astype(str)
                    df_removed_2['FEATUREID'] = df_removed['FEATUREID'].astype(str)
                    df_removed_2['SEGMENTID'] = df_removed['SEGMENTID'].astype(str)
                    df_removed_2['LATITUDE'] = df_removed['LATITUDE'].astype(str)
                    #df_removed_2['ELEVATION'] = df_removed['ELEVATION'].astype(str)
                    df_removed_2['LONGITUDE'] = df_removed['LONGITUDE'].astype(str)

                else:
                    df_removed = df_designreport[df_designreport["FEATUREID"] == "REMOVE"]
                    df_removed_2 = df_removed[
                        ['SEGMENTID',"MILEPOST", "DATASOURCEID", "FEATUREID","LATITUDE","LONGITUDE"]]
                    #['SEGMENTID', "MILEPOST", "DATASOURCEID", "FEATUREID", "LATITUDE", "LONGITUDE", "ELEVATION"]]
                    df_removed_2['MILEPOST'] = df_removed['MILEPOST'].astype(str)
                    df_removed_2['DATASOURCEID'] = df_removed['DATASOURCEID'].astype(str)
                    df_removed_2['FEATUREID'] = df_removed['FEATUREID'].astype(str)
                    df_removed_2['SEGMENTID'] = df_removed['SEGMENTID'].astype(str)
                    df_removed_2['LATITUDE'] = df_removed['LATITUDE'].astype(str)
                    #df_removed_2['ELEVATION'] = df_removed['ELEVATION'].astype(str)
                    df_removed_2['LONGITUDE'] = df_removed['LONGITUDE'].astype(str)
        #df_designremove = df_removed_2.rename(columns={'FEATUREID': 'Status'})
        df_removechanged=pd.concat([df_removed_2, df_changed_2])


        colum_remove = df_removechanged.columns[df_removechanged.isin(['REMOVE']).any()]
        print(colum_remove.values[0])
        df_designremovechanged = df_removechanged.rename(columns={'FEATUREID': 'Status'})

print('------------------------changed data')
print(df_designremovechanged)


#df_designfinal = df_designremovechanged.astype(df_designreporttrackseg.dtypes)
# df_designfinal = df_designreporttrackseg.astype(df_designremovechanged.dtypes)

# filter_data = (df_designreporttrackseg['SEGMENTID']==df_designremovechanged['SEGMENTID'])

df_designremovechanged['MILEPOST']=df_designremovechanged['MILEPOST'].map(str)
df_designremovechanged['DATASOURCEID']=df_designremovechanged['DATASOURCEID'].map(str)
df_designremovechanged['SEGMENTID']=df_designremovechanged['SEGMENTID'].map(str)
df_designremovechanged['LONGITUDE']=df_designremovechanged['LONGITUDE'].map(str)
df_designremovechanged['LATITUDE']=df_designremovechanged['LATITUDE'].map(str)
#df_designremovechanged['ELEVATION']=df_designremovechanged['ELEVATION'].map(str)


df_designreporttracksegment['BEGINMILEPOST']=df_designreporttracksegment['BEGINMILEPOST'].map(str)
df_designreporttracksegment['TRACKNAME']=df_designreporttracksegment['TRACKNAME'].map(str)
df_designreporttracksegment['SEGMENTID']=df_designreporttracksegment['SEGMENTID'].map(str)



df_merged = pd.merge(df_designremovechanged, df_designreporttracksegment, on="SEGMENTID", how="inner")
df_merged.drop(['BEGINMILEPOST'], axis = 1, inplace = True)

print(df_merged)

#Checking with Compare Report data and updating to Compare Report for colours-------------------------------------------------------------\
def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

obj_workbook = openpyxl.load_workbook(file1)
#print(obj_workbook.sheetnames)
obj_sheetm = obj_workbook['Milepost Helper']
print(obj_sheetm)
#print(obj_sheetm)
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)


# This will Color the Changed Data to Orange
max_column = obj_sheetm.max_column
max_row_All = obj_sheetm.max_row
obj_sheet_tables = []
#print(len(final))

# if len(final)>=1:
#     print('-------Enter  RED------------')
#for rows in df_designremovechanged:
for rows in df_merged:
    m = 11  #row 34 =15 and 35 =11-----------------add data dynamically
    f = 10  #header 34 = 14 and 35 = 10
    #row = rows + 1
    #print(row)
    list=[]
    #temp_values_changed = df_designremovechanged[rows]
    temp_values_changed = df_merged[rows]
    print(temp_values_changed)
    #print(temp_values_changed)
    for x, y in temp_values_changed.items():
        x=rows
        list.append(rows)
        if y =='REMOVE':
            y='Removed'
        list_data_list = (str(x) + ':' + str(y))
        for l in range(f, f + 1):
            for j in range(m, row_count+1):
                for k in range(1, max_column + 1):
                    cell_obj = obj_sheetm.cell(row=j, column=k)
                    cell_obj_header = obj_sheetm.cell(row=l, column=k)
                    a = str(list_data_list) #------design data with header
                    cellvalue = cell_obj_header.value
                    celldata = str(cellvalue)
                    b = str(str(cellvalue) + ":" + str(cell_obj.value))#------compare data with header
                    if rows in list and cellvalue in list:
                        # if b.__contains__('LATITUDE'):
                        #     s, e = b.split(':')
                        #     b = ':'.join([s, "{:f}".format(float(e).__round__(6))])
                        # if b.__contains__('LONGITUDE'):
                        #     s, e = b.split(':')
                        #     b = ':'.join([s, "{:f}".format(float(e),6)])
                        # if a!=b:
                        # if a.__eq__(b) :
                        #     cell_obj.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                        if a.__eq__(b):
                            #cell_obj.font = Font(color='00CCFFCC')
                            cell_obj.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                            obj_sheetm.sheet_properties.tabColor = 'CCFFCC'
                        elif a.__ne__(b)  and cell_obj.fill == '00CCFFCC':
                            continue
                            #cell_obj.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        elif a.__ne__(b) and cell_obj.fill.fill_type == None:
                            #cell_obj.font = Font(color='FF0000')
                            cell_obj.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            obj_sheetm.sheet_properties.tabColor = 'FFC7CE'


# if len(final)==0:
#         obj_sheetm.sheet_properties.tabColor = 'CCFFCC'
#
# if len(final)>=1:
#         obj_sheetm.sheet_properties.tabColor = 'FFC7CE'


max_column = obj_sheetm.max_column
max_row = obj_sheetm.max_row
maxRef = [max_row, max_column]
max_row: int
tab = Table(displayName='Table13', ref='A17:{}{}'.format(colnum_string(maxRef[0]), maxRef[1]))

tab.tableStyleInfo = style
obj_workbook.save(file1)

#-----------------------------------------------------ADDED---------------------------------------------------------------------------------
def Split_function(data1):
    data_split = data1.split(":", 1)
    x = data_split[0]
    y = data_split[1]
    return x, y


obj_workbook = openpyxl.load_workbook(file)
print(obj_workbook.sheetnames)
obj_sheet = obj_workbook['Milepost Helper']
df_design_doc = pd.DataFrame(df_designreport)
total_column = len(df_design_doc.columns)
total_row = len(df_design_doc.index)
len(obj_sheet._tables)

style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
obj_sheet = obj_workbook.active
m1 = 2
temarray = []
t_column = 9
z = 1
datalist_added = []
f = 1


for l in range(f, f+1):
    for j in range(m1, total_row + 2):
        cell_obj = obj_sheet.cell(row=j, column=1)
        a = cell_obj.font.color.rgb

        if cell_obj.font.color.rgb == "FF006100":
            Added_list1 = []
            for k2 in range(z, t_column + 1):
                data = obj_sheet.cell(row=j, column=k2).value
                cell_obj_header = obj_sheet.cell(row=l, column=k2)
                cellvalue = cell_obj_header.value
                b = str(str(cellvalue) + ":" + str(data))
                print(b)
                Added_list1.append(b)
            datalist_added.append(Added_list1)
            print('------------------------------------added data----------')
            print(datalist_added)


    # Checking with Compare Report data and updating to Compare Report for colours-------------------------------------------------------------\
    def colnum_string(n):
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string


    obj_workbook = openpyxl.load_workbook(file1)
    # print(obj_workbook.sheetnames)
    obj_sheetm = obj_workbook['Milepost Helper']
    print(obj_sheetm)
    # print(obj_sheetm)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)

    # This will Color the Changed Data to Orange
    max_column = obj_sheetm.max_column
    max_row_All = obj_sheetm.max_row
    obj_sheet_tables = []

m = 15
f = 14
f1 = 14

for a1 in datalist_added:
    for data in a1:
        print('-------------------------------data-------------')
        x3, y3 = Split_function(data)
        a = str(str(x3) + ":" + str(y3))
        #print(a)
        for l in range(f, f + 1):
            for j in range(m, 22):
                for k in range(1, max_column + 1):
                    cell_obj = obj_sheetm.cell(row=j, column=k)
                    cell_obj_header = obj_sheetm.cell(row=l, column=k)
                    #a = str(datalist_added)
                    cellvalue = cell_obj_header.value
                    celldata = str(cellvalue)
                    b = str(str(cellvalue) + ":" + str(cell_obj.value))

                    if b== 'Status:None':
                        for l1 in range(f1, f1 + 1):
                            for k1 in range(1, max_column + 1):
                                cell_obj = obj_sheetm.cell(row=j, column=k1)
                                cell_obj_header = obj_sheetm.cell(row=l1, column=k1)
                                z = str(str(cell_obj_header.value) + ":" + str(cell_obj.value))
                                #print(z)
                                if a.__eq__(z):
                                    #cell_obj.font = Font(color='00CCFFCC')
                                    cell_obj.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                                    obj_sheetm.sheet_properties.tabColor = 'CCFFCC'
                                elif a.__ne__(z)  and cell_obj.fill == '00CCFFCC':
                                    continue
                                    #cell_obj.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                elif a.__ne__(z) and cell_obj.fill.fill_type == None:
                                    #cell_obj.font = Font(color='FF0000')
                                    cell_obj.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                    obj_sheetm.sheet_properties.tabColor = 'FFC7CE'


# if len(final)==0:
#         obj_sheetm.sheet_properties.tabColor = 'CCFFCC'
#
# if len(final)>=1:
#         obj_sheetm.sheet_properties.tabColor = 'FFC7CE'


max_column = obj_sheetm.max_column
max_row = obj_sheetm.max_row
maxRef = [max_row, max_column]
max_row: int
tab = Table(displayName='Table13', ref='A17:{}{}'.format(colnum_string(maxRef[0]), maxRef[1]))

tab.tableStyleInfo = style
obj_workbook.save(file1)